from books import Book
from openpyxl import load_workbook


def pass_hash(
    passwd,
):  # Very basic hashing algorithm just to not store passwords in plain text
    passwd_hash = [str(ord(x)) for x in passwd]
    for i in range(len(passwd_hash)):
        if len(passwd_hash[i]) != 3:
            passwd_hash[i] = "0" * (3 - len(passwd_hash[i])) + passwd_hash[i]
    return "".join(passwd_hash)


def create_book(wb_sheet, i):
    return Book(
        wb_sheet.cell(row=i, column=1).value,
        wb_sheet.cell(row=i, column=2).value,
        wb_sheet.cell(row=i, column=3).value,
        wb_sheet.cell(row=i, column=4).value,
    )


def list_genres():
    wb = load_workbook(filename="lms_books.xlsx")
    wb_sheet = wb.active
    genres = set()
    for i in range(2, wb_sheet.max_row + 1):
        value = str(wb_sheet.cell(row=i, column=4).value)
        if value != "None":
            genres.add(value)

    count = 1
    print("List of Genres:\n")
    for i in genres:
        print(f"{count}. {i.capitalize()}")
        count += 1
