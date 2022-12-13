from books import Book
from populate import populate
from colorama import Fore
from openpyxl import load_workbook


class Shelf:
    def __init__(self, genre):
        self.genre = genre
        self.books = self.populate_books("lms_books.xlsx")
        self.book_count = self.get_book_count()

    def get_book_count(self):
        return len(self.books)

    def populate_books(self, file):
        return populate(self.genre, file)

    @staticmethod
    def show_catalog(title, books):

        print("\n\n\n" + "=" * 136)
        print("|{:^134}|".format(title))
        print("=" * 136)
        # Making a table of books, column headers Name ISBN Author Genre
        print(
            "|{:^5}|{:^55}|{:^20}|{:^30}|{:^20}|".format(
                "Sr No", "Name", "ISBN No.", "Author", "Genre"
            )
        )
        print("=" * 136)
        for i in books:

            print(
                "|{:^5}|{:^55}|{:^20}|{:^30}|{:^20}|".format(
                    books.index(i) + 1, i.name, i.isbn, i.author, i.genre
                )
            )
            print("-" * 136)
        print("\n\n")

    def remove_book(self, user):
        if user.priv == "LIB":
            Shelf.show_catalog(self.genre, self.books)
            if len(self.books) != 0:
                n = int(input("Enter the Sr No. of the book you want to remove: ")) - 1
                if n < self.book_count:
                    Shelf.show_catalog(
                        "Are you sure you want to remove this book?", [self.books[n]]
                    )
                    response = input("(YES/NO)>>")
                    if response.lower() == "yes":
                        wb = load_workbook(filename="lms_books.xlsx")
                        wb_sheet = wb.active
                        for i in range(2, wb_sheet.max_row + 1):
                            if (
                                wb_sheet.cell(row=i, column=2).value
                                == self.books[n].isbn
                            ):
                                wb_sheet.delete_rows(i, 1)
                                wb.save("lms_books.xlsx")
                                break

                    else:
                        print(Fore.RED + "Operation Cancelled" + Fore.white)
                else:
                    print("Enter a valid book no.")
            else:
                print("\nThere are no books in this genre to remove")

        else:
            print(
                "You are not allowed to remove books from a shelf, please check your permissions and try again"
            )

    def add_book(self, user):
        if user.priv == "LIB":
            name = input("Enter the name of the book")
            try:
                isbn = int(input("Enter the isbn number: "))
                author = input("Enter authors name: ")
                book = Book(name, isbn, author, self.genre)
                self.books.append(book)
                wb = load_workbook(filename="lms_books.xlsx")
                wb_sheet = wb.active
                wb_sheet.append(
                    [
                        book.name,
                        book.isbn,
                        book.author,
                        book.genre,
                        "Available",
                        0,
                        0,
                        0,
                    ]
                )
                wb.save("lms_books.xlsx")

            except ValueError:
                print("ISBN number should be an integer")

        else:
            print(
                "You are not allowed to add books to a shelf, please check your permissions and try again"
            )
