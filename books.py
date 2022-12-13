from colorama import Fore
from openpyxl import load_workbook
from datetime import date
import logging


class Book:
    def __init__(self, name, isbn, author, genre):
        self.name = name
        self.author = author
        self.isbn = isbn
        self.genre = genre

    def __str__(self):

        return f"\n\nName: {self.name}\nAuthor: {self.author}\nGenre: {self.genre}\nISBN: {self.isbn}\n"

    def get_status(self):
        wb = load_workbook(filename="lms_books.xlsx")
        wb_sheet = wb.active
        for i in range(2, wb_sheet.max_row + 1):
            if wb_sheet.cell(row=i, column=2).value == self.isbn:
                self.status = wb_sheet.cell(row=i, column=5).value
                date_start = wb_sheet.cell(row=i, column=6).value
                date_end = wb_sheet.cell(row=i, column=7).value
                break
        # This assumes that if a book is issued it cannot be reserved before it is returned
        if self.status == "Issued":
            self.issued_date = date_start
            self.return_date = date_end
            print(Fore.RED + f"Status: Issued {self.return_date}" + Fore.WHITE)

        elif self.status == "Reserved":
            self.reserve_start = date_start
            self.reserve_end = date_end
            print(
                Fore.YELLOW
                + f"Status: Reserved from {self.reserve_start} upto {self.reserve_end}"
                + Fore.WHITE
            )
        elif self.status == "Available":
            print(Fore.GREEN + "Status: Available" + Fore.WHITE)
        wb.close()

    def return_book(self, return_user):
        wb = load_workbook(filename="lms_books.xlsx")
        wb_sheet = wb.active
        for i in range(2, wb_sheet.max_row + 1):
            if wb_sheet.cell(row=i, column=2).value == self.isbn:
                wb_sheet.cell(row=i, column=5).value = "Available"
                wb_sheet.cell(row=i, column=6).value = 0
                wb_sheet.cell(row=i, column=7).value = 0
                wb_sheet.cell(row=i, column=8).value = 0
                break
        print(Fore.GREEN + "\n\nBook Successfully Returned\n\n" + Fore.WHITE)
        wb.save("lms_books.xlsx")
        logging.basicConfig(
            filename="lms.log",
            format="%(asctime)s %(message)s",
            filemode="a",
        )
        logger = logging.getLogger()
        logger.setLevel(logging.DEBUG)

        logger.info(f"{self.name} by {self.author} returned by {return_user}")
        return True

    def issue_book(self, user):
        logging.basicConfig(
            filename="lms.log",
            format="%(asctime)s %(message)s",
            filemode="a",
        )
        logger = logging.getLogger()
        logger.setLevel(logging.DEBUG)

        def enter_issue_date():
            wb_sheet.cell(row=i, column=6).value = date.today()
            try:
                ret_date = [
                    int(x)
                    for x in input("Enter end date in format (yyyy-mm-dd): ").split("-")
                ]
                if ret_date[0] == date.today().year:
                    if ret_date[1] == date.today().month:
                        if ret_date[2] > date.today().day:
                            wb_sheet.cell(row=i, column=7).value = date(
                                ret_date[0], ret_date[1], ret_date[2]
                            )
                            wb_sheet.cell(row=i, column=8).value = user.uname
                            print(
                                Fore.GREEN
                                + "\n\nBook Successfully Issued\n\n"
                                + Fore.WHITE
                            )
                            wb.save("lms_books.xlsx")
                            logger.info(
                                f"{self.name} by {self.author} issued to {user.uname} upto {'-'.join([ str(x) for x in ret_date])}"
                            )
                            return True
                        else:
                            print("\nEnter a valid date")
                            return False
                    elif ret_date[1] > date.today().month:
                        wb_sheet.cell(row=i, column=7).value = date(
                            ret_date[0], ret_date[1], ret_date[2]
                        )
                        wb_sheet.cell(row=i, column=8).value = user.uname
                        print(
                            Fore.GREEN + "\n\nBook Successfully Issued\n\n" + Fore.WHITE
                        )
                        wb.save("lms_books.xlsx")
                        logger.info(
                            f"{self.name} by {self.author} issued to {user.uname} upto {'-'.join([ str(x) for x in ret_date])}"
                        )
                        return True

                    else:
                        print("\nEnter a valid date")
                        return False
                elif ret_date[0] > date.today().year:
                    wb_sheet.cell(row=i, column=7).value = date(
                        ret_date[0], ret_date[1], ret_date[2]
                    )
                    wb_sheet.cell(row=i, column=8).value = user.uname
                    print(Fore.GREEN + "\n\nBook Successfully Issued\n\n" + Fore.WHITE)
                    wb.save("lms_books.xlsx")
                    return True
                else:
                    print("\nEnter a valid date")
                    return False
            except ValueError:
                print("\nEnter a valid date")
                return False

        wb = load_workbook(filename="lms_books.xlsx")
        wb_sheet = wb.active
        for i in range(2, wb_sheet.max_row + 1):
            if wb_sheet.cell(row=i, column=2).value == self.isbn:
                wb_sheet.cell(row=i, column=5).value = "Issued"
                x = enter_issue_date()
                while not x:
                    x = enter_issue_date()
                    continue
                break

    def reserve_book(self, user):
        logging.basicConfig(
            filename="lms.log",
            format="%(asctime)s %(message)s",
            filemode="a",
        )
        logger = logging.getLogger()
        logger.setLevel(logging.DEBUG)

        def enter_reserve_date():
            try:
                res_start_date = [
                    int(x)
                    for x in input("Enter start date in format (yyyy-mm-dd): ").split(
                        "-"
                    )
                ]
                if res_start_date[0] == date.today().year:
                    if res_start_date[1] == date.today().month:
                        if res_start_date[2] > date.today().day:
                            wb_sheet.cell(row=i, column=6).value = date(
                                res_start_date[0], res_start_date[1], res_start_date[2]
                            )
                        else:
                            print("\nEnter a valid date")
                            return False
                    elif res_start_date[1] == date.today().month:
                        wb_sheet.cell(row=i, column=6).value = date(
                            res_start_date[0], res_start_date[1], res_start_date[2]
                        )
                    else:
                        print("\nEnter a valid date")
                        return False
                elif res_start_date[0] > date.today().year:
                    wb_sheet.cell(row=i, column=6).value = date(
                        res_start_date[0], res_start_date[1], res_start_date[2]
                    )

                else:
                    print("\nEnter a valid date")
                    return False

                res_end_date = [
                    int(x)
                    for x in input("Enter end date in format (yyyy-mm-dd): ").split("-")
                ]
                if res_end_date[0] == res_start_date[0]:
                    if res_end_date[1] == res_start_date[1]:
                        if res_end_date[2] > res_start_date[2]:
                            wb_sheet.cell(row=i, column=7).value = date(
                                res_end_date[0], res_end_date[1], res_end_date[2]
                            )
                            wb_sheet.cell(row=i, column=8).value = user.uname
                            print(
                                Fore.GREEN
                                + "\n\nBook Successfully Reserved\n\n"
                                + Fore.WHITE
                            )
                            wb.save("lms_books.xlsx")
                            logger.info(
                                f"{self.name} by {self.author} reserved for {user.uname} from {'-'.join([str(x) for x in res_start_date])} upto {'-'.join([ str(x) for x in res_end_date])}"
                            )
                            return True
                        else:
                            print("\nEnter a valid date")
                            return False
                    elif res_end_date[1] > res_start_date[1]:
                        wb_sheet.cell(row=i, column=7).value = date(
                            res_end_date[0], res_end_date[1], res_end_date[2]
                        )
                        wb_sheet.cell(row=i, column=8).value = user.uname
                        print(
                            Fore.GREEN
                            + "\n\nBook Successfully Reserved\n\n"
                            + Fore.WHITE
                        )
                        wb.save("lms_books.xlsx")
                        logger.info(
                            f"{self.name} by {self.author} reserved for {user.uname} from {'-'.join([str(x) for x in res_start_date])} upto {'-'.join([ str(x) for x in res_end_date])}"
                        )
                        return True
                    else:
                        print("\nEnter a valid date")
                        return False
                elif res_end_date[0] > res_start_date[0]:
                    wb_sheet.cell(row=i, column=7).value = date(
                        res_end_date[0], res_end_date[1], res_end_date[2]
                    )
                    wb_sheet.cell(row=i, column=8).value = user.uname
                    print(
                        Fore.GREEN + "\n\nBook Successfully Reserved\n\n" + Fore.WHITE
                    )
                    wb.save("lms_books.xlsx")
                    logger.info(
                        f"{self.name} by {self.author} reserved for {user.uname} from {'-'.join([str(x) for x in res_start_date])} upto {'-'.join([ str(x) for x in res_end_date])}"
                    )
                    return True
                else:
                    print("\nEnter a valid date")
                    return False
            except ValueError:
                print("\nEnter a valid date")
                return False

        wb = load_workbook(filename="lms_books.xlsx")
        wb_sheet = wb.active
        for i in range(2, wb_sheet.max_row + 1):
            if wb_sheet.cell(row=i, column=2).value == self.isbn:
                wb_sheet.cell(row=i, column=5).value = "Reserved"
                x = enter_reserve_date()
                while not x:
                    x = enter_reserve_date()
                break
