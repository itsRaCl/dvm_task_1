from getpass import getpass
from shelf import Shelf
from openpyxl import load_workbook
from scripts import pass_hash, create_book, list_genres
import datetime

user_dat = {
    "User_1": ["084101115116105110103064049050051", "USR"],
    "Librarian": ["076105098114097114105097110064049050051", "LIB"],
}
# This dictionary can be stored in an sql database too


class User:
    def __init__(self, priv):

        self.uname = input("\nUsername: ")
        if self.uname in user_dat:
            if priv != user_dat[self.uname][1]:
                print("\n\nYou don't have permission to log into this section\n\n")
                exit()
            else:
                self.passwd = pass_hash(getpass("\nPassword: "))
                if self.passwd == user_dat[self.uname][0]:
                    print("Logged In!\n\n")
                    del self.passwd
                else:
                    print("Wrong password\n\nProgram Terminated\n")
                    exit()
        else:
            print("User does not exist contact adimistrators for help")
            exit()


class Basic_User(User):
    def __init__(self):
        self.priv = "USR"
        super().__init__(self.priv)

    def book_actions(self, book_obj):
        book_obj.get_status()
        if book_obj.status != "Available":
            print(
                f"Sorry the book is already {book_obj.status}, Try again after the book is available\n\n"
            )
        else:
            while True:
                try:
                    choice = int(
                        input(
                            "What do you want to do:\n1.Issue Book\n2.Reserve Book\n0. Exit\n>>"
                        )
                    )
                    if choice in [0, 1, 2]:
                        if choice == 1:
                            book_obj.issue_book(self)
                            break
                        if choice == 2:
                            book_obj.reserve_book(self)
                            break
                        if choice == 0:
                            break
                    else:
                        print("Enter a valid choice")
                except ValueError:
                    print("Enter a valid option 0,1,2 from the above choices")

    def options(self):
        while True:
            try:
                choice = int(
                    input(
                        "\n\n1.Search Genre\n2.Search Book via Book Title\n3.Search Book Via ISBN\n4.Search Book Via Author Name\n5.Return Book\n0.Logout\n>>"
                    )
                )
                if choice in range(0, 6):
                    wb = load_workbook(filename="lms_books.xlsx")
                    wb_sheet = wb.active
                    if choice == 1:
                        genres_list = list_genres()
                        genre_shelf = Shelf(
                            genres_list[
                                int(input("\n\nEnter the Genre number: ")) - 1
                            ].lower()
                        )
                        Shelf.show_catalog(genre_shelf.genre, genre_shelf.books)
                        while True:
                            try:
                                book_index = (
                                    int(
                                        input("\nEnter the serial number of the book: ")
                                    )
                                    - 1
                                )
                                if book_index < genre_shelf.book_count:
                                    self.book_actions(genre_shelf.books[book_index])
                                    break
                                else:
                                    print("Enter a valid book number")
                            except ValueError:
                                print("Enter a valid serial number for the book")

                    elif choice == 2:
                        query = input("\nEnter the Book Title: ")
                        for i in range(2, wb_sheet.max_row + 1):
                            cell = wb_sheet.cell(row=i, column=1)
                            if str(cell.value).lower() == query.lower():
                                result_book = create_book(wb_sheet, i)
                                print(result_book)
                                break
                            else:
                                print("The Book is not in the library database")
                        self.book_actions(result_book)
                    elif choice == 3:
                        query = int(input("\nEnter the Book ISBN Number: "))
                        for i in range(2, wb_sheet.max_row + 1):
                            cell = wb_sheet.cell(row=i, column=2)
                            if cell.value == query:
                                result_book = create_book(wb_sheet, i)
                                print(result_book)
                                break
                            else:
                                print("The Book is not in the library database")
                        self.book_actions(result_book)
                    elif choice == 4:
                        query = input("\nEnter the Author's Name: ")
                        books = []
                        for i in range(2, wb_sheet.max_row + 1):
                            cell = wb_sheet.cell(row=i, column=3)
                            if str(cell.value).lower().strip() == query.lower().strip():
                                result_book = create_book(wb_sheet, i)
                                books.append(result_book)
                        if len(books) == 0:
                            print("The Author is not in the database")
                        else:
                            Shelf.show_catalog(query, books)
                            n = int(input("Which book do you want to select:"))
                            self.book_actions(books[n - 1])
                    elif choice == 5:
                        wb = load_workbook(filename="lms_books.xlsx")
                        wb_sheet = wb.active
                        books = []
                        for i in range(2, wb_sheet.max_row + 1):
                            if wb_sheet.cell(row=i, column=8).value == self.uname:
                                books.append(create_book(wb_sheet, i))
                        if len(books) != 0:
                            Shelf.show_catalog(f"Books Issued by {self.uname}", books)

                            choice = (
                                int(
                                    input(
                                        "Enter the Sr No of the book you want to return: "
                                    )
                                )
                                - 1
                            )
                            if choice < len(books):
                                books[choice].return_book(self.uname)
                        else:
                            print("You have no issued books")

                    elif choice == 0:
                        wb.close()
                        break
            except ValueError:
                print("Enter a valid option 0,1,2,3,4,5 from the above choices")


class Librarian(User):
    def __init__(self):
        self.priv = "LIB"
        super().__init__(self.priv)

    def options(self):
        while True:
            try:
                choice = int(
                    input(
                        "Options:\n1.See Issued Books\n2.Manage Books by genre\n0.Logout\n>>"
                    )
                )
                if choice in [0, 1, 2]:
                    if choice == 2:
                        genres_list = list_genres()
                        genre_shelf = Shelf(
                            genres_list[
                                int(
                                    input(
                                        "\n\nEnter the Genre Number you want to manage: "
                                    )
                                )
                                - 1
                            ].lower()
                        )
                        Shelf.show_catalog(
                            genre_shelf.genre + " Books", genre_shelf.books
                        )
                        while True:
                            try:
                                sub_choice = int(
                                    input(
                                        "1.Add Book\n2.Remove Book\n3.See All Books\n0.Exit\n>>"
                                    )
                                )
                                if sub_choice in [0, 1, 2, 3]:
                                    if sub_choice == 2:
                                        genre_shelf.remove_book(self)
                                    if sub_choice == 1:
                                        genre_shelf.add_book(self)
                                    if sub_choice == 0:
                                        break
                                    if sub_choice == 3:
                                        Shelf.show_catalog(
                                            genre_shelf.genre, genre_shelf.books
                                        )
                            except ValueError:
                                print(
                                    "Enter a valid option 0,1,2,3 from the above choices"
                                )

                    elif choice == 1:
                        books = self.get_issued_books()
                        print("\n\n\n" + "=" * 143)
                        print("|{:^141}|".format("Issued Books"))
                        print("=" * 143)
                        # Making a table of books, column headers Name ISBN Author Genre
                        print(
                            "|{:^5}|{:^50}|{:^20}|{:^30}|{:^10}|{:^11}|{:^9}|".format(
                                "Sr No",
                                "Name",
                                "ISBN No.",
                                "Author",
                                "Genre",
                                "Issued Upto",
                                "Issued By",
                            )
                        )
                        print("=" * 143)
                        for i in books:
                            print(
                                "|{:^5}|{:^50}|{:^20}|{:^40}|{:^10}|{:^11}|{:^9}|".format(
                                    books.index(i) + 1,
                                    i[0].name,
                                    i[0].isbn,
                                    i[0].author,
                                    i[0].genre,
                                    "-".join(
                                        [str(i[1].day), str(i[1].month), str(i[1].year)]
                                    ),
                                    i[2],
                                )
                            )
                            print("-" * 113)
                        print("\n\n")
                    elif choice == 0:
                        print("Logging Out!\n\nBye!\n\n")
                        break
            except ValueError:
                print("Enter a valid option 0,1,2 from the above choices")

    def get_issued_books(self):
        books = []
        wb = load_workbook(filename="lms_books.xlsx")
        wb_sheet = wb.active
        for i in range(2, wb_sheet.max_row + 1):
            if wb_sheet.cell(row=i, column=5).value == "Issued":
                books.append(
                    [
                        create_book(wb_sheet, i),
                        wb_sheet.cell(row=i, column=7).value,
                        wb_sheet.cell(row=i, column=8).value,
                    ]
                )
        return books
