from openpyxl import load_workbook
from books import Book


def populate(genre, file):
    books = []
    wb = load_workbook(filename=file)
    wb_sheet = wb.active
    for i in range(2, wb_sheet.max_row + 1):
        if wb_sheet.cell(row=i, column=4).value == genre:
            result_book = Book(
                wb_sheet.cell(row=i, column=1).value,
                wb_sheet.cell(row=i, column=2).value,
                wb_sheet.cell(row=i, column=3).value,
                genre,
            )
            books.append(result_book)
    return books
